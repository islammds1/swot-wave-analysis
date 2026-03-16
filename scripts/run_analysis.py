"""
scripts/run_analysis.py
=======================
Orchestration script for the SWOT HR PIXC wave cross-spectrum analysis.

Usage
-----
    python scripts/run_analysis.py

All scene-specific parameters (file paths, box definitions, spectral
settings) are in config/boxes_D1_D2_D3.py — edit that file, not this one.

Output
------
  • swot_hr_boxes_phase.png          — diagnostic spectra (E_ssh / Coh / Phase)
  • swot_hr_pub_spectra_phase.jpg    — publication-quality figure
  • swot_wave_results_<date>_...xlsx — full results spreadsheet
  • swot_wave_results_<date>_....csv — same data as CSV
"""

# ── stdlib ────────────────────────────────────────────────────────────────────
import os
import re
import sys
import functools

# ── third-party ───────────────────────────────────────────────────────────────
import numpy as np
import xarray as xr
import matplotlib
import matplotlib.pyplot as plt
import pandas as pd
from pyproj import Proj

# ── local src ─────────────────────────────────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.swot_geometry import get_swot_geometry
from src.spectral      import bin_to_sat_grid, welch_cross_spectrum
from src.ambiguity     import (resolve_180_ardhuin2024,
                                sat_frame_to_geo_north,
                                circular_stats, ci95_circular)
from src.bathymetry    import mean_depth_for_box
from src.dispersion    import period_from_k

from config.boxes_D1_D2_D3 import (
    PIXC_FILE, PIXC_FILE2, LR_FILE, BATHY_FILE, GROUP,
    BOXES, BOX_LABELS, BOX_TILE,
    RES, LWIN, LAMBDA_MIN, LAMBDA_MAX, G,
    LOOK_SIDE, SWATH,
    ARD_DK_REL, ARD_DTHETA, ARD_COH_MIN,
    COH_MIN_PATCH, OUT_DIR,
)

# ============================================================
# 1) SATELLITE GEOMETRY
# ============================================================

geo            = get_swot_geometry(LR_FILE, swath=SWATH)
trackangle     = geo["trackangle_deg"]
pass_direction = geo["pass_direction"]
Vsat_E, Vsat_N = geo["Vsat_EN"]

print(f"Track angle : {trackangle:.2f}°  ({pass_direction})")


def utm_to_sat(dx, dy):
    """Rotate UTM offset (East, North) into satellite frame (along, cross)."""
    return dx * Vsat_E + dy * Vsat_N,  dx * Vsat_N - dy * Vsat_E


def sat_to_utm(u_a, u_c):
    """Inverse rotation: satellite frame → UTM offset."""
    return u_a * Vsat_E + u_c * Vsat_N,  u_a * Vsat_N - u_c * Vsat_E


def box_corners_utm(cx, cy, half_a, half_c):
    sa = np.array([-half_a,  half_a,  half_a, -half_a])
    sc = np.array([-half_c, -half_c,  half_c,  half_c])
    dx, dy = sat_to_utm(sa, sc)
    return np.column_stack([cx + dx, cy + dy])


# ── Partially apply geometry to the injected callables ───────────────────────

def _resolve(ka_cand, kc_cand, C_patch, Ka, Kc, coh_patch):
    return resolve_180_ardhuin2024(
        ka_cand, kc_cand, C_patch, Ka, Kc, coh_patch,
        look_side=LOOK_SIDE,
        dk_rel=ARD_DK_REL,
        dtheta_win=ARD_DTHETA,
        coh_min=ARD_COH_MIN,
    )


def _geo_dir(ka, kc):
    return sat_frame_to_geo_north(ka, kc, Vsat_E, Vsat_N)


# ============================================================
# 2) LOAD PIXC  (both tiles)
# ============================================================

proj = Proj(proj="utm", zone=30, ellps="WGS84")


def _load_pixc(fpath):
    ds      = xr.open_dataset(fpath, group=GROUP, engine="h5netcdf")
    lat     = ds["latitude"].values;   lon  = ds["longitude"].values
    ssh     = ds["height"].values;     sig0 = ds["sig0"].values
    good    = np.isfinite(sig0) & (sig0 > 0)
    sig0_db = np.full_like(sig0, np.nan, dtype=float)
    sig0_db[good] = 10.0 * np.log10(sig0[good])
    ok = np.isfinite(lat) & np.isfinite(lon) & np.isfinite(ssh) & np.isfinite(sig0_db)
    x, y = proj(lon[ok], lat[ok])
    return x, y, ssh[ok], sig0_db[ok]


x1, y1, ssh1, sig1 = _load_pixc(PIXC_FILE)
x2, y2, ssh2, sig2 = _load_pixc(PIXC_FILE2)

x_utm    = np.concatenate([x1, x2])
y_utm    = np.concatenate([y1, y2])
ssh_all  = np.concatenate([ssh1, ssh2])
sig0_all = np.concatenate([sig1, sig2])
tile_tag = np.concatenate([np.zeros(len(x1), dtype=np.int8),
                            np.ones (len(x2), dtype=np.int8)])

print(f"Tile 0 (067R): {x1.size:,} pts  |  Tile 1 (065R): {x2.size:,} pts")

# ============================================================
# 3) LOAD BATHYMETRY
# ============================================================

bathy_ds  = xr.open_dataset(BATHY_FILE)
elevation = bathy_ds["elevation"]
_lat_name = "lat" if "lat" in elevation.dims else "latitude"
_lon_name = "lon" if "lon" in elevation.dims else "longitude"

# ============================================================
# 4) MAIN LOOP OVER BOXES
# ============================================================

results = []

for i_box, (cx, cy, half_a, half_c) in enumerate(BOXES):
    label = BOX_LABELS[i_box]
    print(f"\n{'='*60}")
    print(f"{label}  centre ({cx/1e3:.1f}, {cy/1e3:.1f}) km  "
          f"size {2*half_a/1e3:.1f}×{2*half_c/1e3:.1f} km")

    # ── Route to correct tile ─────────────────────────────────────────────────
    t  = BOX_TILE[i_box]
    xp = x_utm[tile_tag == t];  yp = y_utm[tile_tag == t]
    sp = ssh_all[tile_tag == t]; gp = sig0_all[tile_tag == t]

    # ── Bin onto satellite-frame grid ─────────────────────────────────────────
    ssh_g, sig_g, _, _, n_pts = bin_to_sat_grid(
        xp, yp, sp, gp, cx, cy, half_a, half_c, RES, utm_to_sat)

    if ssh_g is None:
        print("  ⚠ Not enough points — skipping.")
        results.append(None);  continue

    ny, nx = ssh_g.shape
    pct    = 100.0 * np.sum(np.isfinite(ssh_g)) / (ny * nx)
    print(f"  Points: {n_pts}  Grid: {ny}×{nx}  valid: {pct:.0f}%")

    # ── Mean depth for this box ───────────────────────────────────────────────
    h_box = mean_depth_for_box(cx, cy, half_a, half_c, elevation,
                                utm_to_sat, proj,
                                lat_name=_lat_name, lon_name=_lon_name)
    print(f"  Mean depth: {h_box:.1f} m" if np.isfinite(h_box)
          else "  Mean depth: no bathymetry data")

    # ── Welch cross-spectrum ──────────────────────────────────────────────────
    out = welch_cross_spectrum(
        ssh_g, sig_g,
        res=RES, lwin=LWIN,
        lam_min=LAMBDA_MIN, lam_max=LAMBDA_MAX,
        resolve_fn=_resolve,
        geo_dir_fn=_geo_dir,
        g=G, h=h_box,
        coh_min_patch=COH_MIN_PATCH,
    )

    if out is None:
        print("  ⚠ No valid Welch patches — skipping.")
        results.append(None);  continue

    n_pat = out["n_patches"]
    print(f"  Welch patches: {n_pat}")

    # ── Wavelength statistics ─────────────────────────────────────────────────
    fin_K  = out["patch_K"][np.isfinite(out["patch_K"])]
    n_lam  = len(fin_K)

    peak_lam      = float(2*np.pi / np.median(fin_K))         if n_lam > 0 else np.nan
    peak_std_lam  = float(np.std(2*np.pi / fin_K, ddof=1))    if n_lam > 1 else np.nan
    peak_ci95_lam = 1.96 * peak_std_lam / np.sqrt(n_lam)      if n_lam > 1 else np.nan

    # ── Direction statistics (circular) ──────────────────────────────────────
    fin_dir = out["patch_dir_geo"][np.isfinite(out["patch_dir_geo"])]
    peak_dir_geo, peak_std_dir, R, n_dir = circular_stats(fin_dir)
    peak_ci95_dir = ci95_circular(peak_std_dir, n_dir)

    peak_dir_sat = (float(np.nanmedian(
        out["patch_dir_sat"][np.isfinite(out["patch_dir_sat"])]))
        if np.any(np.isfinite(out["patch_dir_sat"])) else np.nan)

    # ── Peak period with box-specific depth ───────────────────────────────────
    if n_lam > 0 and np.isfinite(h_box) and h_box > 0:
        T_peak, regime = period_from_k(float(np.median(fin_K)), h_box, G)
    else:
        T_peak, regime = np.nan, "unknown"

    print(f"  λ = {peak_lam:.1f} ± {peak_std_lam:.1f} m  "
          f"T = {T_peak:.2f} s  ({regime})")
    print(f"  θ = {peak_dir_geo:.1f} ± {peak_std_dir:.1f}°  "
          f"R = {R:.2f}  n_dir = {n_dir}")

    # ── Directional reliability flag ──────────────────────────────────────────
    # Thresholds follow the uncertainty framework in the manuscript:
    #   R >= 0.95  → high quality       → 95% CI ≈ ±2–5°
    #   R >= 0.83  → moderate quality   → 95% CI ≈ ±10–14°
    #   R >= 0.70  → low quality        → use with caution
    #   R <  0.70  → unreliable         → excluded from quantitative analysis
    if R < 0.70:
        dir_quality  = "unreliable"
        dir_reliable = False
        print(f"  ⚠ WARNING: R = {R:.2f} < 0.70 — directional estimate "
              f"flagged as UNRELIABLE for {label}. "
              f"Excluded from quantitative refraction/diffraction analysis.")
    elif R < 0.83:
        dir_quality  = "low"
        dir_reliable = True
        print(f"  ⚠ NOTE: R = {R:.2f} — low directional confidence for {label}. "
              f"Use with caution.")
    elif R < 0.95:
        dir_quality  = "moderate"
        dir_reliable = True
        print(f"  ℹ  R = {R:.2f} — moderate directional confidence for {label}.")
    else:
        dir_quality  = "high"
        dir_reliable = True
        print(f"  ✓  R = {R:.2f} — high directional confidence for {label}.")

    results.append({
        "label"         : label,
        "box_utm"       : (cx, cy, half_a, half_c),
        "box_idx"       : i_box,
        **out,
        # Wavelength
        "peak_lam"      : peak_lam,
        "peak_std_lam"  : peak_std_lam,
        "peak_ci95_lam" : peak_ci95_lam,
        "n_lam_patches" : n_lam,
        # Direction
        "peak_dir_geo"  : peak_dir_geo,
        "peak_std_dir"  : peak_std_dir,
        "peak_ci95_dir" : peak_ci95_dir,
        "peak_dir_sat"  : peak_dir_sat,
        "dir_R"         : R,
        "n_dir_patches" : n_dir,
        # Directional reliability
        "dir_reliable"  : dir_reliable,
        "dir_quality"   : dir_quality,
        # Depth & period
        "h_mean"        : h_box,
        "T_peak"        : T_peak,
        "T_regime"      : regime,
    })

print(f"\n✓ {sum(r is not None for r in results)}/{len(BOXES)} boxes processed.")

# ============================================================
# 5) SUMMARY TABLE
# ============================================================

print(f"\n{'='*90}")
print(f"{'Box':<8} {'h [m]':>7} {'λ [m]':>8} {'±1σ':>6} {'95CI':>6} "
      f"{'T [s]':>6} {'θ [°N]':>8} {'±1σ':>6} {'R':>5} "
      f"{'Quality':>12} {'Regime':>12}")
print("-"*90)
for r in results:
    if r is None:
        continue
    flag = "  ← EXCLUDED" if not r["dir_reliable"] else ""
    print(f"{r['label']:<8} {r['h_mean']:>7.1f} {r['peak_lam']:>8.1f} "
          f"{r['peak_std_lam']:>6.1f} {r['peak_ci95_lam']:>6.1f} "
          f"{r['T_peak']:>6.2f} {r['peak_dir_geo']:>8.1f} "
          f"{r['peak_std_dir']:>6.1f} {r['dir_R']:>5.2f} "
          f"{r['dir_quality']:>12} {r['T_regime']:>12}{flag}")

# ============================================================
# 6) DIAGNOSTIC PLOTS — E_ssh | Coherence | Phase
# ============================================================

box_colors = plt.cm.tab10(np.linspace(0, 0.9, max(len(BOXES), 1)))
N_BOXES    = len(BOXES)
KLIM       = 6.0

fig, axes = plt.subplots(3, N_BOXES, figsize=(3.5*N_BOXES, 11),
                          sharex=True, sharey=True)
if N_BOXES == 1:
    axes = axes[:, np.newaxis]

for i, r in enumerate(results):
    ax0, ax1, ax2 = axes[0, i], axes[1, i], axes[2, i]
    if r is None:
        ax0.set_title(f"{BOX_LABELS[i]}\nno data", fontsize=9)
        for ax in (ax0, ax1, ax2): ax.axis("off")
        continue

    ka = r["k_along"] / (2*np.pi) * 1e3
    kc = r["k_cross"] / (2*np.pi) * 1e3

    im0 = ax0.pcolormesh(kc, ka, np.log10(r["E_ssh"] + 1e-12),
                          cmap="viridis", shading="auto")
    ax0.set_title(f"{r['label']}\nlog₁₀(E_ssh)", fontsize=9)
    ax0.set_xlim(-KLIM, KLIM); ax0.set_ylim(-KLIM, KLIM)
    fig.colorbar(im0, ax=ax0, pad=0.02)

    im1 = ax1.pcolormesh(kc, ka, r["coherence"],
                          cmap="plasma", vmin=0, vmax=1, shading="auto")
    ax1.set_title("Coherence SSH–σ₀", fontsize=9)
    ax1.set_xlim(-KLIM, KLIM); ax1.set_ylim(-KLIM, KLIM)
    fig.colorbar(im1, ax=ax1, pad=0.02)

    im2 = ax2.pcolormesh(kc, ka, r["phase_deg"],
                          cmap="seismic", vmin=-180, vmax=180, shading="auto")

    # ── Colour-code phase panel background by directional reliability ─────────
    if r["dir_R"] < 0.70:
        ax2.set_facecolor("#ffe6e6")   # light red  → unreliable
        reliability_note = f"\n⚠ R={r['dir_R']:.2f} UNRELIABLE"
        title_color = "red"
    elif r["dir_R"] < 0.83:
        ax2.set_facecolor("#fff3e0")   # light orange → low
        reliability_note = f"\n(low R={r['dir_R']:.2f})"
        title_color = "darkorange"
    elif r["dir_R"] < 0.95:
        ax2.set_facecolor("#fffde7")   # light yellow → moderate
        reliability_note = f"\n(mod R={r['dir_R']:.2f})"
        title_color = "goldenrod"
    else:
        reliability_note = f"\n(R={r['dir_R']:.2f} ✓)"
        title_color = "darkgreen"

    ax2.set_title(
        f"Phase C_cross [°]\n"
        f"λ={r['peak_lam']:.0f}±{r['peak_std_lam']:.0f} m  "
        f"θ={r['peak_dir_geo']:.0f}±{r['peak_std_dir']:.0f}°"
        f"{reliability_note}",
        fontsize=7, color=title_color)
    ax2.set_xlim(-KLIM, KLIM); ax2.set_ylim(-KLIM, KLIM)
    cb2 = fig.colorbar(im2, ax=ax2, pad=0.02)
    cb2.set_ticks([-180, -90, 0, 90, 180])

for row in range(3):
    axes[row, 0].set_ylabel("k_along [cycles/km]", fontsize=9)
fig.text(0.5, -0.01, "k_cross [cycles/km]", ha="center", va="top", fontsize=11)
plt.suptitle(f"SWOT HR  |  track {trackangle:.1f}° ({pass_direction})  "
             f"| {LAMBDA_MIN}–{LAMBDA_MAX} m  | Ardhuin 2024", fontsize=11)
plt.tight_layout()
plt.savefig(os.path.join(OUT_DIR, "swot_hr_boxes_phase.png"),
            dpi=150, bbox_inches="tight")
plt.show()

# ============================================================
# 7) PUBLICATION PLOT
# ============================================================

PUB_LABELS   = ["D1A", "D1D", "D2", "D3A", "D3D"]
KLIM_PUB     = 8.0
res_by_label = {r["label"]: r for r in results if r is not None}

plt.rcParams.update({
    "font.family": "Times New Roman", "font.size": 24,
    "font.weight": "bold", "axes.labelweight": "bold",
    "axes.titleweight": "bold",
})

n_pub    = len(PUB_LABELS)
fig_pub, axes_pub = plt.subplots(3, n_pub, figsize=(3.2*n_pub, 10),
                                  sharex=True, sharey=True)
ROW_TITLES = [r"$\log_{10}(E_{ssh})$", r"Coherence",
              r"Phase of $C_{SSH,\sigma_0}$ [°]"]
CMAPS      = ["viridis", "plasma", "seismic"]
VLIMS      = [None, (0, 1), (-180, 180)]

for col, lbl in enumerate(PUB_LABELS):
    r = res_by_label.get(lbl)
    for row in range(3):
        ax = axes_pub[row, col]
        if r is None:
            ax.text(0.5, 0.5, f"{lbl}\nno data", ha="center", va="center",
                    transform=ax.transAxes, fontsize=14, fontweight="bold")
            ax.set_xlim(-KLIM_PUB, KLIM_PUB); ax.set_ylim(-KLIM_PUB, KLIM_PUB)
            continue
        ka = r["k_along"] / (2*np.pi) * 1e3
        kc = r["k_cross"] / (2*np.pi) * 1e3
        data = ([np.log10(r["E_ssh"] + 1e-12), r["coherence"], r["phase_deg"]])[row]
        vmin = VLIMS[row][0] if VLIMS[row] else None
        vmax = VLIMS[row][1] if VLIMS[row] else None
        im   = ax.pcolormesh(kc, ka, data, cmap=CMAPS[row], shading="auto",
                              vmin=vmin, vmax=vmax)
        ax.set_xlim(-KLIM_PUB, KLIM_PUB); ax.set_ylim(-KLIM_PUB, KLIM_PUB)

        if row == 0:
            # ── Add reliability badge to column title in pub figure ───────────
            if r["dir_quality"] == "unreliable":
                badge = " ⚠"
                tc    = "red"
            elif r["dir_quality"] == "low":
                badge = " !"
                tc    = "darkorange"
            else:
                badge = ""
                tc    = "black"
            ax.set_title(lbl + badge, fontsize=16, fontweight="bold",
                         pad=6, color=tc)

        if col == n_pub - 1:
            cb = fig_pub.colorbar(im, ax=ax, pad=0.03, fraction=0.046)
            cb.ax.tick_params(labelsize=16)
            if row == 2: cb.set_ticks([-180, -90, 0, 90, 180])
            cb.set_label(ROW_TITLES[row], fontsize=16, fontweight="bold")
        ax.tick_params(labelsize=16)

sx = fig_pub.supxlabel(r"$k_{cross}$ [cycles km$^{-1}$]",
                        fontsize=24, fontweight="bold")
sy = fig_pub.supylabel(r"$k_{along}$ [cycles km$^{-1}$]",
                        fontsize=24, fontweight="bold")
sx.set_y(0.09); sy.set_x(0.09)
plt.tight_layout(rect=[0.04, 0.04, 1.0, 1.0])
plt.savefig(os.path.join(OUT_DIR, "swot_hr_pub_spectra_phase.jpg"),
            dpi=600, bbox_inches="tight", pad_inches=0.09)
plt.show()
plt.rcParams.update(matplotlib.rcParamsDefault)

# ============================================================
# 8) EXPORT TO EXCEL + CSV
# ============================================================

_fname  = os.path.basename(PIXC_FILE)
_parts  = _fname.split("_")
try:
    _cycle    = _parts[4]
    _pass     = _parts[5]
    _tile_1   = _parts[6]
    _date_str = _parts[7][:8]
    acq_date  = pd.to_datetime(_date_str, format="%Y%m%d").date()
except Exception:
    _cycle = _pass = _tile_1 = "unknown";  acq_date = None;  _date_str = "unknown"

_tile_2 = os.path.basename(PIXC_FILE2).split("_")[6]
_tile_name_map = {0: _tile_1, 1: _tile_2}


def _zone(lbl):
    m = re.match(r"^([A-Z]+\d+)", lbl)
    return m.group(1) if m else lbl


records = []
for i_b, r in enumerate(results):
    if r is None:
        continue
    if r["label"] == "D2":
        continue
    cx, cy, half_a, half_c = r["box_utm"]
    lam_ci95 = r.get("peak_ci95_lam", np.nan)

    def _f(v):
        return round(float(v), 4) if np.isfinite(v) else np.nan

    records.append({
        "date"              : acq_date,
        "cycle"             : _cycle,
        "pass"              : _pass,
        "box_id"            : r["label"],
        "zone"              : _zone(r["label"]),
        "tile"              : _tile_name_map.get(BOX_TILE[i_b], "unknown"),
        "cx_utm_m"          : cx,
        "cy_utm_m"          : cy,
        "half_along_m"      : half_a,
        "half_cross_m"      : half_c,
        "depth_m"           : _f(r.get("h_mean",  np.nan)),
        "wavelength_m"      : _f(r["peak_lam"]),
        "wavelength_1sig_m" : _f(r["peak_std_lam"]),
        "wavelength_95ci_m" : _f(lam_ci95),
        "period_s"          : _f(r.get("T_peak",  np.nan)),
        "dispersion_regime" : r.get("T_regime", "unknown"),
        "direction_degN"    : _f(r["peak_dir_geo"]),
        "direction_1sig"    : _f(r["peak_std_dir"]),
        "direction_95ci"    : _f(r["peak_ci95_dir"]),
        "dir_R"             : _f(r["dir_R"]),
        # ── Reliability columns ───────────────────────────────────────────────
        # dir_quality  : "high" (R≥0.95) | "moderate" (R≥0.83) |
        #                "low"  (R≥0.70) | "unreliable" (R<0.70)
        # dir_reliable : True if R≥0.70; False if R<0.70
        #                Boxes with dir_reliable=False are excluded from the
        #                quantitative refraction / diffraction analysis in the
        #                manuscript (D3C and D3D are the expected candidates).
        "dir_quality"       : r.get("dir_quality",  "unknown"),
        "dir_reliable"      : r.get("dir_reliable", False),
        "n_patches"         : r["n_patches"],
        "n_lam_patches"     : r["n_lam_patches"],
        "n_dir_patches"     : r["n_dir_patches"],
        "track_angle_deg"   : round(trackangle, 3),
        "pass_direction"    : pass_direction,
    })

df   = pd.DataFrame(records)
stem = f"swot_wave_results_{_date_str}_cycle{_cycle}_pass{_pass}"

# ── Report unreliable boxes at end of run ─────────────────────────────────────
df_unreliable = df[df["dir_reliable"] == False].copy()
if not df_unreliable.empty:
    print(f"\n{'='*60}")
    print("⚠ Boxes flagged as directionally UNRELIABLE (R < 0.70):")
    print(df_unreliable[["box_id", "dir_R", "dir_quality"]].to_string(index=False))
    print("  → Direction values retained in output file but flagged.")
    print("  → Exclude these boxes from quantitative refraction analysis.")
    print(f"{'='*60}")
else:
    print("\n✓ All boxes passed the directional reliability threshold (R ≥ 0.70).")

xlsx_path = os.path.join(OUT_DIR, stem + ".xlsx")
csv_path  = os.path.join(OUT_DIR, stem + ".csv")

with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="SWOT_wave_results")
    ws = writer.sheets["SWOT_wave_results"]

    # ── Highlight unreliable rows in red in the Excel output ─────────────────
    from openpyxl.styles import PatternFill, Font
    red_fill   = PatternFill(start_color="FFE6E6", end_color="FFE6E6",
                              fill_type="solid")
    orange_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0",
                               fill_type="solid")
    red_font   = Font(color="CC0000", bold=True)

    # Find column index for dir_quality (1-based in openpyxl)
    header_row  = [cell.value for cell in ws[1]]
    try:
        q_col = header_row.index("dir_quality") + 1
        r_col = header_row.index("dir_reliable") + 1
    except ValueError:
        q_col = r_col = None

    if q_col is not None:
        for row_idx, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
            quality_val  = ws.cell(row=row_idx, column=q_col).value
            reliable_val = ws.cell(row=row_idx, column=r_col).value
            if quality_val == "unreliable":
                for cell in row_cells:
                    cell.fill = red_fill
                ws.cell(row=row_idx, column=q_col).font = red_font
            elif quality_val == "low":
                for cell in row_cells:
                    cell.fill = orange_fill

    # ── Auto-size columns ─────────────────────────────────────────────────────
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = max_len + 3

df.to_csv(csv_path, index=False, float_format="%.4f")

print(f"\n✓ Saved:  {xlsx_path}")
print(f"✓ Saved:  {csv_path}")
print(f"  {len(df)} rows  |  {len(df['box_id'].unique())} boxes")
print(f"  Reliable boxes : {df['dir_reliable'].sum()}")
print(f"  Unreliable boxes (R<0.70, excluded from analysis): "
      f"{(~df['dir_reliable']).sum()}")
print("\n" + df.to_string(index=False))